from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import openpyxl

# Initialize driver
driver = webdriver.Chrome()

# Login
driver.get("http://wrts.lge.com/wrts_main_page.php")
driver.maximize_window()
time.sleep(5)
driver.find_element(By.ID, "username").send_keys("nithyanantham.s")
driver.find_element(By.ID, "password").send_keys("LGsoft@1234567")
driver.find_element(By.XPATH, "//input[@type='submit']").click()

# Handle optional popup
try:
    close_btn = WebDriverWait(driver, 30).until(
        EC.element_to_be_clickable((By.ID, "id_shut_down_checkbox"))
    )
    close_btn.click()
except:
    pass

# Setup Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Test Results"
headers = ['S.No', 'Region', 'WRTS ID', 'Total', 'Total NA', 'Pass', 'Fail', 'NA', 'NE', 'NI', 'Blocked']
ws.append(headers)

region_info = [
    ("EU", "89974"),
    ("US", "89975"),
    ("BR", "89977"),
    ("TW", "89979"),
    ("PH", "89978"),
    ("AJ", "89976"),
    ("Panel", "89980")
]

sno = 1

for region, wrts_id in region_info:
    try:
        webID = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((By.XPATH, f"//td[@title='{wrts_id}']"))
        )
        webID.click()
        print(f"{region} - {wrts_id} Clicked")
    except Exception as e:
        print(f"{region} - {wrts_id} Not Found: {e}")
        continue

    # Click model status button
    driver.find_element(By.ID, "btn_main_view_model_status").click()

    # Wait for new window and switch
    WebDriverWait(driver, 30).until(EC.number_of_windows_to_be(2))
    windows = driver.window_handles
    driver.switch_to.window(windows[1])

    # Wait for table
    try:
        WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, "//table[@class='ui-jqgrid-ftable']"))
        )
    except:
        print(f"[{region}] Table not found initially.")
        driver.close()
        driver.switch_to.window(windows[0])
        continue

    time.sleep(2)
    driver.execute_script("document.querySelector('.ui-dialog-titlebar-close').click();")

    # Retry logic to fetch rows
    max_retries = 3
    retry_delay = 3
    rows = []

    for attempt in range(max_retries):
        rows = driver.find_elements(By.CSS_SELECTOR, "table.ui-jqgrid-ftable tbody tr")
        if rows:
            break
        print(f"[{region}] Attempt {attempt+1} failed — no rows found. Retrying...")
        time.sleep(retry_delay)
        driver.refresh()
        WebDriverWait(driver, 30).until(
            EC.visibility_of_element_located((By.XPATH, "//table[@class='ui-jqgrid-ftable']"))
        )
        driver.execute_script("document.querySelector('.ui-dialog-titlebar-close').click();")

    if not rows:
        print(f"[{region}] Failed to fetch data after {max_retries} retries.")
        driver.close()
        driver.switch_to.window(windows[0])
        continue

    for row in rows:
        cells = row.find_elements(By.TAG_NAME, 'td')
        if len(cells) >= 12:
            values = []
            for i in range(4, 12):
                text = cells[i].text.strip().replace(',', '')
                if text == '':
                    values.append(0)
                else:
                    try:
                        num = int(text) if text.isdigit() else float(text)
                        values.append(num)
                    except ValueError:
                        values.append(0)

            data = [sno, region, wrts_id] + values
            ws.append(data)
            sno += 1
            break  # stop after the first valid data row

    driver.close()
    driver.switch_to.window(windows[0])
    time.sleep(1)

# Save Excel file
wb.save("wrts_all_regions.xlsx")
print("Excel file saved.")

# Calculate totals
total_columns = [0] * 8
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=11):
    for i, cell in enumerate(row):
        try:
            value = int(str(cell.value).replace(',', ''))
            total_columns[i] += value
        except:
            continue

grand_total_row = ["", "", "Grand Total"] + total_columns
ws.append(grand_total_row)

wb.save("wrts_all_regions_with_total.xlsx")
print("Excel file with Grand Total saved.")

driver.quit()
